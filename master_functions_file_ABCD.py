# -*- coding: utf-8 -*-
"""
Created on Tue Aug  2 15:25:23 2022

@author: juarez
"""
import pandas as pd
import time, os, sys
import openpyxl


# =============================================================================
# The following function helps to obtain all the conditons for the test
# It is up to the user to make sure it matches its intention
# The function returns  tuple which cannot be modified while running the program
# it is imperative that all conditions are th ones needed to avoid double work
# =============================================================================
def extract_settings_from_conditions_file(settings_file):
    #Generating Placeholder variable for conditions
    test_freq = ""
    test_volt = ""
    test_usid = ""
    test_cond = ""
    test_path = ""
    real_number = ""
    imaginary_number = ""
    
    reading_conditions_file = openpyxl.load_workbook(settings_file)
    sheet_file_to_work_with = reading_conditions_file.active
    
    #Get the boundaries of files, we know in advance how much memory is needed
    last_row_containing_data_in_file = sheet_file_to_work_with.max_row
    
    for each_row_in_file in range(1, last_row_containing_data_in_file+1):
        cell_content = sheet_file_to_work_with.cell(row = each_row_in_file, column = 1)
        cell_content = cell_content.value
        if str(cell_content).lower() == "freq (mhz)":
            test_freq = sheet_file_to_work_with.cell(row = each_row_in_file, column = 2)
            test_freq = test_freq.value
        if str(cell_content).lower() == "vio (v)":
            test_volt = sheet_file_to_work_with.cell(row = each_row_in_file, column = 2)
            test_volt = test_volt.value
        if str(cell_content).lower() == "usid config":
            test_usid = sheet_file_to_work_with.cell(row = each_row_in_file, column = 2)
            test_usid = test_usid.value
        if str(cell_content).lower() == "test condition":
            test_cond = sheet_file_to_work_with.cell(row = each_row_in_file, column = 2)
            test_cond = test_cond.value
        if str(cell_content).lower() == "test path":
            test_path = sheet_file_to_work_with.cell(row = each_row_in_file, column = 2)
            test_path = test_path.value 
            
        getting_port_value = sheet_file_to_work_with.cell(row = each_row_in_file, column = 2)
        getting_port_value = getting_port_value.value
        
        if str(getting_port_value) == str(test_path):
            real_number = sheet_file_to_work_with.cell(row = each_row_in_file, column = 3)
            real_number = real_number.value
            imaginary_number = sheet_file_to_work_with.cell(row = each_row_in_file, column = 4)
            imaginary_number = imaginary_number.value
    return test_freq, test_volt, test_usid, test_cond, test_path, real_number, imaginary_number


# =============================================================================
# The following function extracts data from the csv file
# =============================================================================

def get_csv_file_to_extract_data_conditions(file_to_convert_into_data_frame):
    # global pow_incident, third_harmonics, ivio_at_beginning, ivio_before_breakdown
    
    pow_incident = ""
    third_harmonics = ""
    ivio_at_beginning = ""
    ivio_before_breakdown = ""
    
    file_to_processed = pd.read_csv(file_to_convert_into_data_frame)
    data_top = file_to_processed.keys()
    
    #Getting Incident Power one before breakdown
    try:    
        pow_incident = file_to_processed.iloc[[-2], [2]]
        pow_incident = pow_incident.iloc[0,0]
    except:
        pass
    #Get 3rd Harmonics one before break down
    try:
        third_harmonics = file_to_processed.iloc[[-2], [6]]
        third_harmonics = third_harmonics.iloc[0,0]
    except:
        pass
    #Get IVIO at the bginning of test
    try:
       ivio_at_beginning = file_to_processed.iloc[[0], [1]]
       ivio_at_beginning = ivio_at_beginning.iloc[0,0]
    except:
        pass
    #Get IVIO at the one before break down
    try:
        ivio_before_breakdown = file_to_processed.iloc[[-2], [1]]
        ivio_before_breakdown = ivio_before_breakdown.iloc[0,0]
    except:
        pass
    
    if pow_incident == "" or third_harmonics == "" or ivio_at_beginning == "" or ivio_before_breakdown == "":
        return None, None, None, None
        # return None, None, None, None
    else:
        return pow_incident, third_harmonics, ivio_at_beginning, ivio_before_breakdown
    



# =============================================================================
# The following function is design to extract the condition, ABCD SP file 
# and RFBD Data File from a data matrix file
# =============================================================================
def build_data_container_from_matrix_file(data_matrix_file):
    reading_matrix_file = openpyxl.load_workbook(data_matrix_file)
    reading_matrix_worksheet = reading_matrix_file.active
    
    master_data_list_for_matrix_file = []
    for each_row in reading_matrix_worksheet.iter_rows():
        building_temp_list =[]
        for cell in each_row:
            building_temp_list.append(cell.value)
        
        master_data_list_for_matrix_file.append(building_temp_list)   
    
    return master_data_list_for_matrix_file

# data_matrix_file = r"C:\Users\juarez\OneDrive - Qualcomm\Desktop\ABCD_Method_tool\dat_matrix_file_template.xlsx"
# get_data_container = build_data_container_from_matrix_file(data_matrix_file)
# print(get_data_container)






# file_to_process = r"C:\Users\juarez\OneDrive - Qualcomm\Desktop\ABCD_Method_tool\QAT_data_files\QAT5616_V200_B2ATE_RFBD\QAT5616_V200_B2ATE_EVB_900MHz_ALL_OFF_Active_RFc_SN001_2022-07-19_22_33.csv"
# # file_to_process = r"C:\Users\juarez\OneDrive - Qualcomm\Desktop\ABCD_Method_tool\QAT_data_files\QAT5616_V200_B2ATE_RFBD\QAT5616_V200_B2ATE_EVB_900MHz_ALL_OFF_Active_RFc_SN002_2022-07-19_22_49.csv"
# # file_to_process = r"C:\Users\juarez\OneDrive - Qualcomm\Desktop\ABCD_Method_tool\QAT_data_files\QAT5616_V200_B2ATE_RFBD\QAT5616_V200_B2ATE_EVB_900MHz_ALL_OFF_Active_RFc_SN007_2022-07-20_02_59.csv"
# pow_incident, third_harmonics, ivio_at_beginning, ivio_before_breakdown = get_csv_file_to_extract_data_conditions(file_to_process)


# print(pow_incident)
# print(third_harmonics)
# print(ivio_at_beginning)
# print(ivio_before_breakdown)